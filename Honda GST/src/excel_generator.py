import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


def _safe_float(value):
    if value is None:
        return 0.0

    if isinstance(value, (int, float)):
        return float(value)

    value = str(value).strip().replace(",", "")
    if value in ("", "-", "nan", "None"):
        return 0.0

    return float(value)


def generate_excel(df, dealer_code):
    base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    output_dir = os.path.join(base_dir, "output", "attachments")
    os.makedirs(output_dir, exist_ok=True)

    file_path = os.path.join(output_dir, f"{dealer_code}.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = "Invoice Details"

    headers = [
        "Dealer Code",
        "Particulars",
        "Invoice Date",
        "Invoice",
        "Taxable",
        "IGST",
        "CGST",
        "SGST",
    ]

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    start_data_row = 2

    for row_num, (_, row) in enumerate(df.iterrows(), start=start_data_row):
        values = [
            row["Dealer code"],
            row["Particulars"],
            row["Invoice Date"],
            row["Invoice"],
            _safe_float(row["Taxable"]),
            _safe_float(row["IGST"]),
            _safe_float(row["CGST"]),
            _safe_float(row["SGST"]),
        ]

        for col_num, value in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.border = border
            cell.alignment = left if col_num < 5 else center

            if col_num >= 5:
                cell.number_format = '#,##0.00'

    total_row = ws.max_row + 1
    ws.cell(row=total_row, column=4, value="Total").font = Font(bold=True)

    for col_num in range(4, 9):
        ws.cell(row=total_row, column=col_num).border = border
        ws.cell(row=total_row, column=col_num).font = Font(bold=True)

    for col_num in range(5, 9):
        col_letter = get_column_letter(col_num)
        ws.cell(
            row=total_row,
            column=col_num,
            value=f"=SUM({col_letter}{start_data_row}:{col_letter}{total_row - 1})",
        )
        ws.cell(row=total_row, column=col_num).number_format = '#,##0.00'

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:H{ws.max_row}"

    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max_length + 3

    wb.save(file_path)
    return file_path